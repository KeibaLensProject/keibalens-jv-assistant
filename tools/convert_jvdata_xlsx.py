"""JV-Data仕様書_4.9.0.1.xlsx -> Markdown 変換器。

出力 (references/spec/jv-data-spec/):
  README.md                  表紙 + 目次
  change-history.md          変更履歴
  special-notes.md           特記事項
  data-types.md              データ種別一覧
  data-timing.md             データ提供タイミング･提供単位
  record-formats/            フォーマット (レコード種別ごとに分割) + README
  code-tables/               コード表 (コードごとに分割) + README

使い方: python tools/convert_jvdata_xlsx.py
"""
import io
import os
import re
import sys
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SPEC = os.path.join(ROOT, "plugins", "jra-van-spec-skills", "skills",
                    "jra-van-spec-skills", "references", "spec")
SRC = os.path.join(ROOT, "context", "JV-Data仕様書_4.9.0.1.xlsx")
OUT = os.path.join(SPEC, "jv-data-spec")
VERSION = "4.9.0.1"

Z2H = str.maketrans("０１２３４５６７８９", "0123456789")


def full2half(s):
    return s.translate(Z2H)


def trimrow(row):
    vals = ["" if c is None else str(c).strip() for c in row]
    last = max([j for j, v in enumerate(vals) if v], default=-1)
    return vals[: last + 1]


def cell(s):
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("|", "\\|").replace("\n", "<br>")
    return s.strip()


def safe_name(s):
    s = re.sub(r'[<>:"/\\|?*]', "", s)
    s = s.replace("（", "(").replace("）", ")")
    return s.strip().strip(".")


def write(path, text):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(text)
    print("  wrote", os.path.relpath(path, ROOT), f"({len(text)} chars)")


def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        r = list(r) + [""] * (len(headers) - len(r))
        out.append("| " + " | ".join(cell(x) for x in r[: len(headers)]) + " |")
    return "\n".join(out)


# --------------------------------------------------------------------------
# フォーマット -> record-formats/
# --------------------------------------------------------------------------
def parse_reclen(v):
    if "レコード長" in v:
        idx = v.index("レコード長")
        for j in range(idx + 1, len(v)):
            if v[j].replace(",", "").isdigit():
                return v[j]
    return ""


def convert_format(ws):
    rows = [trimrow(r) for r in ws.iter_rows(values_only=True)]
    is_sec = lambda v: len(v) > 1 and re.match(r"^[0-9０-９]+．", v[1])

    intro, records, cur = [], [], None
    for v in rows:
        if is_sec(v):
            m = re.match(r"^([0-9０-９]+)．(.*)$", v[1])
            cur = {
                "num": int(full2half(m.group(1))),
                "name": m.group(2).strip(),
                "reclen": parse_reclen(v),
                "header": None,
                "notes": [],
                "rows": [],
            }
            records.append(cur)
        elif cur is None:
            if any(v):
                intro.append(v)
        else:
            if len(v) > 1 and v[1] == "項番":
                cur["header"] = v
            elif cur["header"] is None:
                if any(v):
                    cur["notes"].append(v)
            else:
                if any(v):
                    cur["rows"].append(v)

    os.makedirs(os.path.join(OUT, "record-formats"), exist_ok=True)
    index = []
    for rec in records:
        recid = ""
        for v in rec["rows"]:
            name = v[4] if len(v) > 4 else ""
            if name == "レコード種別ID":
                desc = v[10] if len(v) > 10 else ""
                mm = re.search(r'"([0-9A-Za-z]{2})"', desc)
                if mm:
                    recid = mm.group(1).upper()
                break
        rec["recid"] = recid
        fname = f"{rec['num']:03d}-{recid or 'XX'}-{safe_name(rec['name'])}.md"
        rec["file"] = fname

        hdr = rec["header"] or []
        matrix = hdr[11:] if len(hdr) > 11 else []
        headers = ["項番", "キー", "項目名", "位置", "繰返", "バイト", "合計",
                   "初期値", "説明"] + [m for m in matrix]

        # 位置(数値)を持つ行が項目行。持たない行は脚注(凡例・注記)として分離
        field_rows, foot_rows = [], []
        for v in rec["rows"]:
            pos = v[5] if len(v) > 5 else ""
            if re.match(r"^\(?\s*\d", pos):
                field_rows.append(v)
            else:
                foot_rows.append(v)

        trows = []
        for v in field_rows:
            g = lambda i: v[i] if i < len(v) else ""
            ban = g(1) or g(2)
            base = [ban, g(3), g(4), g(5), g(6), g(7), g(8), g(9), g(10)]
            mat = [g(11 + k) for k in range(len(matrix))]
            trows.append(base + mat)

        foot_txts = [" ".join(x for x in f if x).strip() for f in foot_rows]
        foot_txts = [t for t in foot_txts if t]
        has_legend = any("このデータ区分時に値を設定" in t for t in foot_txts)

        title = f"{rec['num']}．{rec['name']}"
        body = [f"# {title}", ""]
        meta = []
        if recid:
            meta.append(f"- **レコード種別ID**: `{recid}`")
        if rec["reclen"]:
            meta.append(f"- **レコード長**: {rec['reclen']} バイト")
        meta.append(f"- 出典: JV-Data仕様書 {VERSION}「フォーマット」")
        body += meta + [""]
        for n in rec["notes"]:
            txt = " ".join(x for x in n if x).strip()
            if txt:
                body.append("> " + cell(txt))
        if rec["notes"]:
            body.append("")
        if matrix and not has_legend:
            body.append('> 末尾の列(`' + "`,`".join(matrix) +
                        '`)はデータ区分別の収録有無。'
                        '○=このデータ区分時に値を設定 / △=値を設定する場合としない場合が混在 / -=初期値を設定。')
            body.append("")
        body.append(md_table(headers, trows))
        body.append("")
        for t in foot_txts:
            body.append("> " + cell(t))
        if foot_txts:
            body.append("")
        write(os.path.join(OUT, "record-formats", fname), "\n".join(body))
        index.append(rec)

    # record-formats README
    lines = ["# レコードフォーマット一覧", "",
             f"JV-Data仕様書 {VERSION}「フォーマット」シートのレコード定義です。",
             "レコード種別ごとにファイルを分割しています。", ""]
    for v in intro:
        txt = " ".join(x for x in v if x).strip()
        if txt and not txt.startswith("■"):
            lines.append("> " + cell(txt))
    lines += ["", "| No | レコード種別ID | レコード名 | レコード長(byte) | ファイル |",
              "|---|---|---|---|---|"]
    for rec in index:
        lines.append(
            f"| {rec['num']} | {rec['recid'] or ''} | {rec['name']} | "
            f"{rec['reclen']} | [{rec['file']}](./{rec['file']}) |")
    lines.append("")
    write(os.path.join(OUT, "record-formats", "README.md"), "\n".join(lines))
    return index


# --------------------------------------------------------------------------
# コード表 -> code-tables/
# --------------------------------------------------------------------------
def convert_codes(ws):
    rows = [trimrow(r) for r in ws.iter_rows(values_only=True)]
    is_sec = lambda v: len(v) > 1 and re.match(r"^\d{3,4}\.", v[1])

    intro, sections, cur = [], [], None
    for v in rows:
        if is_sec(v):
            cur = {"title": v[1], "rows": []}
            sections.append(cur)
        elif cur is None:
            if any(v):
                intro.append(v)
        else:
            cur["rows"].append(v)

    os.makedirs(os.path.join(OUT, "code-tables"), exist_ok=True)
    index = []
    for sec in sections:
        body_rows = [v for v in sec["rows"] if any(v)]
        content_labels = ["内容"]
        data = []
        bytesize = ""
        hi = None
        for k, v in enumerate(body_rows):
            if len(v) > 2 and v[1] == "バイト数" and v[2] == "値":
                hi = k
                break
        start = 0
        if hi is not None:
            start = hi + 1
            if start < len(body_rows):
                sv = body_rows[start]
                if len(sv) > 3 and not sv[1] and not sv[2] and sv[3]:
                    content_labels = [x for x in sv[3:] if x] or ["内容"]
                    start += 1
        notes = []
        for v in body_rows[start:]:
            val = v[2] if len(v) > 2 else ""
            if len(v) > 1 and v[1] and not bytesize:
                bytesize = v[1]
            if val == "":
                txt = " ".join(x for x in v if x).strip()
                if txt:
                    notes.append(txt)
                continue
            content = [v[3 + i] if 3 + i < len(v) else "" for i in range(len(content_labels))]
            data.append([val] + content)

        num = sec["title"].split(".")[0]
        nm = sec["title"].split(".", 1)[1] if "." in sec["title"] else sec["title"]
        fname = f"{num}-{safe_name(nm)}.md"
        sec["file"] = fname
        sec["num"] = num
        sec["name"] = nm

        headers = ["値"] + content_labels
        out = [f"# {sec['title']}", ""]
        if bytesize:
            out.append(f"- **バイト数**: {bytesize}")
        out.append(f"- 出典: JV-Data仕様書 {VERSION}「コード表」")
        out.append("")
        out.append(md_table(headers, data))
        out.append("")
        for nt in notes:
            out.append("> " + cell(nt))
        if notes:
            out.append("")
        write(os.path.join(OUT, "code-tables", fname), "\n".join(out))
        index.append(sec)

    lines = ["# コード表一覧", "",
             f"JV-Data仕様書 {VERSION}「コード表」シートの各種コード定義です。", ""]
    for v in intro:
        txt = " ".join(x for x in v if x).strip()
        if txt and not txt.startswith("■"):
            lines.append("> " + cell(txt))
    lines += ["", "| コード番号 | 名称 | ファイル |", "|---|---|---|"]
    for sec in index:
        lines.append(f"| {sec['num']} | {sec['name']} | [{sec['file']}](./{sec['file']}) |")
    lines.append("")
    write(os.path.join(OUT, "code-tables", "README.md"), "\n".join(lines))
    return index


# --------------------------------------------------------------------------
# 汎用シート (プロローグ文 + テーブル)
# --------------------------------------------------------------------------
def convert_generic(ws, title, outfile, ffill=0):
    rows = [trimrow(r) for r in ws.iter_rows(values_only=True)]
    out = [f"# {title}", "",
           f"出典: JV-Data仕様書 {VERSION}「{ws.title}」", ""]
    seg = []
    last = [""] * 16

    def flush():
        if not seg:
            return
        width = max(len(r) for r in seg)
        cols = [i for i in range(width)
                if any((r[i] if i < len(r) else "") for r in seg)]
        headers = [f"col{j+1}" for j in range(len(cols))]
        data = [[(r[i] if i < len(r) else "") for i in cols] for r in seg]
        out.append(md_table(headers, data))
        out.append("")
        seg.clear()

    for v in rows:
        if not any(v):
            continue
        text_cells = [x for x in v if x]
        first = text_cells[0]
        if len(text_cells) == 1 and re.match(r"^[（(]\s*[0-9０-９]+[）)]", first):
            flush()
            out.append(f"## {cell(first)}")
            out.append("")
            last = [""] * 16
            continue
        if len(text_cells) == 1 and v[0] == "" and not seg:
            if not first.startswith("■"):
                out.append(cell(first))
                out.append("")
            continue
        row = v[:]
        for i in range(min(ffill, len(row))):
            if row[i] == "":
                row[i] = last[i]
            else:
                last[i] = row[i]
        seg.append(row)
    flush()
    write(os.path.join(OUT, outfile), "\n".join(out))


# --------------------------------------------------------------------------
def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    print("== record-formats ==")
    recs = convert_format(wb["フォーマット"])
    print("== code-tables ==")
    codes = convert_codes(wb["コード表"])
    print("== other sheets ==")
    convert_generic(wb["変更履歴"], f"JVData 変更履歴 (Ver.{VERSION})", "change-history.md", ffill=3)
    convert_generic(wb["特記事項"], "JVData 特記事項", "special-notes.md")
    convert_generic(wb["データ種別一覧"], "JVData データ種別一覧", "data-types.md", ffill=3)
    convert_generic(wb["データ提供タイミング･提供単位"], "JVData データ提供タイミング･提供単位", "data-timing.md", ffill=3)

    lines = [
        f"# JRA-VAN Data Lab. JVData 仕様書 (Ver.{VERSION})", "",
        "本ディレクトリは `context/JV-Data仕様書_4.9.0.1.xlsx` を Markdown 化したものです。",
        "発行: JRAシステムサービス株式会社 / 更新日 2024年8月7日 / 適用日 2024年8月7日", "",
        "## 目次", "",
        "| 区分 | 内容 |",
        "|---|---|",
        f"| [レコードフォーマット](./record-formats/README.md) | 全{len(recs)}レコード種別の項目定義 |",
        f"| [コード表](./code-tables/README.md) | 全{len(codes)}種のコード定義 |",
        "| [変更履歴](./change-history.md) | 仕様の変更履歴 |",
        "| [特記事項](./special-notes.md) | データに関する注意事項 |",
        "| [データ種別一覧](./data-types.md) | データ種別IDとレコード種別の対応 |",
        "| [データ提供タイミング･提供単位](./data-timing.md) | 各データの提供タイミング |",
        "",
        "## JVData の文字コード", "",
        "- 全角文字: Shift_JIS (2バイト)",
        "- 半角文字(英･数･半角ｶﾅ): JIS8 (1バイト)",
        "- レコードの収録順序に仕様はなく、予告なく変わる場合がある。"
        "DB反映時はキー項目やデータ作成日を利用すること。",
        "",
    ]
    write(os.path.join(OUT, "README.md"), "\n".join(lines))
    print("done:", len(recs), "records,", len(codes), "code tables")


if __name__ == "__main__":
    main()
